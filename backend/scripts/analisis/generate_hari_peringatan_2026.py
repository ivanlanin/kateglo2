"""Generate a support workbook of 2026 commemorations from multiple sources."""

from __future__ import annotations

import re
import warnings
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable
from urllib.parse import quote

warnings.filterwarnings(
    "ignore",
    message=r"urllib3 .* doesn't match a supported version",
    category=Warning,
)

import requests
from openpyxl import Workbook
from openpyxl.styles import Font

YEAR = 2026
ROOT_DIR = Path(__file__).resolve().parents[3]
OUTPUT_PATH = ROOT_DIR / ".data" / "pendukung" / "hari-peringatan.xlsx"
WIKIPEDIA_PAGE_URL = "https://id.wikipedia.org/wiki/Daftar_hari_penting_di_Indonesia"
WIKIPEDIA_API_URL = "https://id.wikipedia.org/w/api.php"
NAGER_API_URL = f"https://date.nager.at/api/v3/PublicHolidays/{YEAR}/ID"
USER_AGENT = "Kateglo/2.0 data-support generator"

MONTHS = {
    "Januari": 1,
    "Februari": 2,
    "Maret": 3,
    "April": 4,
    "Mei": 5,
    "Juni": 6,
    "Juli": 7,
    "Agustus": 8,
    "September": 9,
    "Oktober": 10,
    "November": 11,
    "Desember": 12,
}

WIKIPEDIA_BLACKLIST = (
    "Hari Ulang Tahun",
    "Milad",
    "Harlah",
    "Hubungan Diplomatik",
    "Provinsi",
    "Kota ",
    "Kabupaten",
    "Presiden",
    "Wakil Presiden",
    "Perdana Menteri",
    "Pemilihan kepala daerah",
    "Telkom",
    "Indosat",
    "Telkomsel",
    "TVRI",
    "RRI",
    "Pos Indonesia",
)


@dataclass(frozen=True)
class CommemorationRow:
    tanggal: date
    peringatan: str
    kategori: str
    sumber: str
    tautan: str


CURATED_ROWS = [
    CommemorationRow(date(YEAR, 2, 21), "Hari Bahasa Ibu Internasional", "peringatan_internasional", "UNESCO", "https://www.unesco.org/en/days/mother-language"),
    CommemorationRow(date(YEAR, 3, 21), "Hari Puisi Sedunia", "peringatan_internasional", "UNESCO", "https://www.unesco.org/en/days/poetry"),
    CommemorationRow(date(YEAR, 4, 7), "Hari Kesehatan Sedunia", "peringatan_internasional", "WHO", "https://www.who.int/campaigns/world-health-day"),
    CommemorationRow(date(YEAR, 9, 8), "Hari Literasi Internasional", "peringatan_internasional", "UNESCO", "https://www.unesco.org/en/days/literacy"),
    CommemorationRow(date(YEAR, 9, 30), "Hari Penerjemahan Internasional", "peringatan_internasional", "PBB", "https://www.un.org/en/observances/translation-day"),
    CommemorationRow(date(YEAR, 10, 5), "Hari Guru Sedunia", "peringatan_internasional", "UNESCO", "https://www.unesco.org/en/days/teachers"),
    CommemorationRow(date(YEAR, 12, 10), "Hari Hak Asasi Manusia", "peringatan_internasional", "PBB", "https://www.un.org/en/observances/human-rights-day"),
]


def fetch_json(url: str, params: dict | None = None) -> dict | list:
    response = requests.get(
        url,
        params=params,
        timeout=30,
        headers={"User-Agent": USER_AGENT},
    )
    response.raise_for_status()
    return response.json()


def clean_wikitext(text: str) -> str:
    cleaned = re.sub(r"<ref[^>/]*/\s*>", "", text)
    cleaned = re.sub(r"<ref[^>]*>.*?</ref>", "", cleaned)
    cleaned = re.sub(r"\{\{[^{}]*\}\}", "", cleaned)
    cleaned = re.sub(r"\[\[(?:[^\]|]+\|)?([^\]]+)\]\]", r"\1", cleaned)
    cleaned = re.sub(r"\[https?://[^\s\]]+\s+([^\]]+)\]", r"\1", cleaned)
    cleaned = re.sub(r"\[https?://[^\]]+\]", "", cleaned)
    cleaned = re.sub(r"\[\d+\]", "", cleaned)
    cleaned = cleaned.replace("''", "")
    cleaned = cleaned.replace("&nbsp;", " ")
    cleaned = cleaned.replace("–", "-")
    cleaned = cleaned.replace("—", "-")
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip(" :'\"")


def extract_wikipedia_link(raw_item: str) -> str:
    match = re.search(r"\[\[([^\]|#]+)", raw_item)
    if not match:
        return WIKIPEDIA_PAGE_URL

    target = match.group(1).strip()
    if ":" in target:
        return WIKIPEDIA_PAGE_URL

    return "https://id.wikipedia.org/wiki/" + quote(target.replace(" ", "_"), safe=":_()/")


def classify_wikipedia_entry(name: str) -> str:
    international_keywords = (
        "Internasional",
        "Sedunia",
        "Dunia",
        "Global",
        "ASEAN",
        "PBB",
        "UNESCO",
        "WHO",
    )
    national_keywords = (
        "Nasional",
        "Indonesia",
        "Republik Indonesia",
        "Kemerdekaan",
        "Pancasila",
        "Santri",
        "Batik",
        "Sumpah Pemuda",
        "Kartini",
        "Ibu",
        "Bela Negara",
    )

    if any(keyword in name for keyword in international_keywords):
        return "peringatan_internasional"
    if any(keyword in name for keyword in national_keywords):
        return "peringatan_nasional"
    return "peringatan_tematik"


def parse_nager_rows() -> list[CommemorationRow]:
    data = fetch_json(NAGER_API_URL)
    rows = []
    for item in data:
        rows.append(
            CommemorationRow(
                tanggal=date.fromisoformat(item["date"]),
                peringatan=item["localName"],
                kategori="libur_nasional",
                sumber="Nager.Date",
                tautan=NAGER_API_URL,
            )
        )
    return rows


def parse_wikipedia_rows() -> list[CommemorationRow]:
    params = {
        "action": "query",
        "prop": "revisions",
        "titles": "Daftar_hari_penting_di_Indonesia",
        "rvslots": "main",
        "rvprop": "content",
        "format": "json",
        "formatversion": "2",
    }
    data = fetch_json(WIKIPEDIA_API_URL, params=params)
    text = data["query"]["pages"][0]["revisions"][0]["slots"]["main"]["content"]
    body = text.split("== Hari penting lainnya ==", 1)[1].split("== Lihat pula ==", 1)[0]

    current_month = None
    rows = []
    month_pattern = re.compile(
        r"^===\s+(Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September|Oktober|November|Desember)\s+===$"
    )
    date_pattern = re.compile(
        r"^(\d{1,2})\s+(Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September|Oktober|November|Desember)$"
    )

    for raw_line in body.splitlines():
        line = raw_line.strip()
        month_match = month_pattern.match(line)
        if month_match:
            current_month = month_match.group(1)
            continue

        if not current_month or not line.startswith("*"):
            continue

        raw_item = line.lstrip("*").strip()
        plain_item = clean_wikitext(raw_item)
        prefix, separator, rest = plain_item.partition(":")
        if not separator:
            continue

        if "Tanggal tidak diketahui" in prefix or prefix.startswith("(Setiap "):
            continue
        if "-" in prefix:
            continue

        date_match = date_pattern.fullmatch(prefix.strip())
        if not date_match:
            continue

        plain_rest = clean_wikitext(rest)
        if not any(token in plain_rest for token in ("Hari", "Peringatan", "Pekan")):
            continue
        if any(token in plain_rest for token in WIKIPEDIA_BLACKLIST):
            continue

        rows.append(
            CommemorationRow(
                tanggal=date(YEAR, MONTHS[date_match.group(2)], int(date_match.group(1))),
                peringatan=plain_rest,
                kategori=classify_wikipedia_entry(plain_rest),
                sumber="Wikipedia Indonesia",
                tautan=extract_wikipedia_link(raw_item),
            )
        )

    return rows


def dedupe_rows(rows: Iterable[CommemorationRow]) -> list[CommemorationRow]:
    priority = {
        "PBB": 0,
        "UNESCO": 0,
        "WHO": 0,
        "Nager.Date": 1,
        "Wikipedia Indonesia": 2,
    }
    deduped = {}

    for row in rows:
        key = (row.tanggal.isoformat(), row.peringatan.casefold())
        current = deduped.get(key)
        if current is None or priority.get(row.sumber, 99) < priority.get(current.sumber, 99):
            deduped[key] = row

    return sorted(deduped.values(), key=lambda item: (item.tanggal, item.peringatan.casefold(), item.sumber.casefold()))


def build_workbook(rows: list[CommemorationRow]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"hari_peringatan_{YEAR}"
    headers = ["tanggal", "peringatan", "kategori", "sumber", "tautan"]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append([row.tanggal, row.peringatan, row.kategori, row.sumber, row.tautan])
        link_cell = sheet.cell(row=sheet.max_row, column=5)
        link_cell.hyperlink = row.tautan
        link_cell.style = "Hyperlink"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row=row_index, column=1).number_format = "yyyy-mm-dd"

    widths = {
        "A": 14,
        "B": 52,
        "C": 26,
        "D": 20,
        "E": 72,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    return workbook


def main() -> None:
    rows = dedupe_rows([*CURATED_ROWS, *parse_nager_rows(), *parse_wikipedia_rows()])
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook(rows)
    workbook.save(OUTPUT_PATH)
    print(f"Generated {len(rows)} rows -> {OUTPUT_PATH}")


if __name__ == "__main__":
    main()